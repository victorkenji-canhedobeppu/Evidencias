# src/core/excel_reader.py

import pandas as pd
import openpyxl
from datetime import datetime


class Excel_Reader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None

    def _load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            print(f"Erro fatal ao carregar o workbook com openpyxl: {e}")
            self.workbook = None

    def _get_header_type(self) -> str:
        if not self.sheet:
            return None
        has_merged_cells_on_row_13 = any(
            cell_range.min_row == 13 or cell_range.max_row == 13
            for cell_range in self.sheet.merged_cells.ranges
        )
        return "multilevel" if has_merged_cells_on_row_13 else "single"

    def _get_clean_headers(self, header_type: str) -> list:
        if not self.sheet:
            return None
        if header_type == "multilevel":
            headers_l1_raw = [cell.value for cell in self.sheet[13]]
            headers_l2 = [cell.value for cell in self.sheet[14]]
            combined_headers = []
            num_columns = max(len(headers_l1_raw), len(headers_l2))
            for i in range(num_columns):
                h1 = headers_l1_raw[i] if i < len(headers_l1_raw) else None
                h2 = headers_l2[i] if i < len(headers_l2) else None
                final_h1 = h1 or ""
                final_h2 = h2 or ""
                if final_h1:
                    combined_headers.append(f"{final_h1} - {final_h2}".strip(" -"))
                else:
                    combined_headers.append(final_h2)
            return combined_headers
        else:  # 'single'
            return [cell.value for cell in self.sheet[14]]

    def get_data_as_dataframe(self, date_column_name: str) -> dict[str, pd.DataFrame]:
        self._load_workbook()
        if not self.workbook:
            return {}

        all_sheets_data = {}

        for sheet_name in self.workbook.sheetnames:
            print(f"--- Processando a página: {sheet_name} ---")
            self.sheet = self.workbook[sheet_name]

            if self.sheet.max_row < 15:
                print(f"Página '{sheet_name}' ignorada por ter poucas linhas.")
                continue

            header_type = self._get_header_type()
            if not header_type:
                continue

            clean_headers = self._get_clean_headers(header_type)
            if not clean_headers:
                continue

            # --- NOVA E DEFINITIVA LÓGICA DE LEITURA DE DADOS ---
            all_rows_data = []
            # Itera sobre as linhas da planilha, começando da linha 15
            for row in self.sheet.iter_rows(min_row=15):
                processed_row = []
                for cell in row:
                    # Verifica se o valor da célula é um objeto de data do Python
                    if isinstance(cell.value, datetime):
                        # Se for, formata para o nosso padrão de string
                        processed_row.append(cell.value.strftime("%d/%m/%Y"))
                    else:
                        # Se não, apenas converte para string, tratando células vazias (None)
                        processed_row.append(
                            str(cell.value) if cell.value is not None else ""
                        )
                all_rows_data.append(processed_row)

            if not all_rows_data:
                continue

            # Cria o DataFrame a partir dos dados já processados e formatados
            df_sheet = pd.DataFrame(all_rows_data, columns=clean_headers)

            # --- A LÓGICA DE LIMPEZA CONTINUA A MESMA ---
            df_sheet = df_sheet[~(df_sheet == "").all(axis=1)]
            if df_sheet.empty:
                continue

            is_empty_col = (df_sheet == "").all()
            if is_empty_col.any():
                cols_to_drop = is_empty_col[is_empty_col].index
                df_sheet.drop(columns=cols_to_drop, inplace=True)

            is_duplicated = df_sheet.columns.duplicated(keep="first")
            if is_duplicated.any():
                cols_to_drop = df_sheet.columns[is_duplicated]
                df_sheet = df_sheet.loc[:, ~is_duplicated]

            all_sheets_data[sheet_name] = df_sheet.reset_index(drop=True)

        return all_sheets_data
